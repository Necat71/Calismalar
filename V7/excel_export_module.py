

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import QMessageBox
import os
import datetime
import traceback # Hata izini yazdırmak için eklendi


def _generate_excel_output(dialog_instance=None, project_name="", excel_data=None, parent_window=None):
    """
    Proje verilerini Excel olarak dışa aktarır.
    :param dialog_instance: ProjeYonetimDialog nesnesi. İçindeki tableWidget kullanılacak.
    :param project_name: Projenin adı.
    """
    try:
        # dialog_instance'ın geçerli olup olmadığını ve tableWidget özelliğine sahip olup olmadığını kontrol edin
        if dialog_instance is None or not hasattr(dialog_instance, 'tableWidget'):
            QMessageBox.critical(
                None,
                "Hata",
                "Geçersiz diyalog örneği veya tableWidget bulunamadı."
            )
            return

        # Gerçek QTableWidget nesnesini dialog_instance'dan alın
        table_widget = dialog_instance.tableWidget

        # Proje adını dialog_instance.lineEdit_projeadi'dan alın
        # Eğer lineEdit_projeadi yoksa veya boşsa, bir uyarı göster
        if hasattr(dialog_instance, 'lineEdit_projeadi') and dialog_instance.lineEdit_projeadi.text():
            project_name = dialog_instance.lineEdit_projeadi.text()
        
        if not project_name:
            QMessageBox.warning(
                None,
                "Excel Hatası",
                "Teklif formu oluşturmak için önce bir proje kaydetmeli veya proje adı girmelisiniz.",
            )
            return

        if table_widget.rowCount() == 0:
            QMessageBox.information(
                None, "Bilgi", "Teklif formu oluşturmak için tabloda veri bulunmuyor."
            )
            return

        output_directory = r"C:\Users\necat\OneDrive\Belgeler"  # Varsayılan klasör, isteğe göre değiştirilebilir

        if not os.path.exists(output_directory):
            os.makedirs(output_directory)

        today_date = datetime.date.today().strftime("%d.%m.%Y")
        excel_filename = os.path.join(
            output_directory, f"Proje_Teklif_Formu_{project_name}_{today_date}.xlsx" # Proje adını dosya adına ekle
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Proje Teklif Formu"

        # Başlık stilini tanımla
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Başlık ekle
        # Başlık hücrelerini Excel çıktısındaki yeni sütun sayısına göre birleştir
        # Excel'deki son sütun 'N' olacak (14 sütun)
        sheet.merge_cells("A1:N1") 
        title_cell = sheet["A1"]
        title_cell.value = f"Proje Teklif Formu - {project_name} - {today_date}" # Proje adını başlığa ekle
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 30

        # Boş bir satır bırak
        sheet.append([])

        # Tablo başlıkları (UI tablosundakiyle eşleşen ve Excel'de gösterilecekler)
        # UI tablosundaki sütun indeksleri ile Excel başlıklarını hizalayalım
        excel_headers = [
            "Sıra No",           # Excel col 1 (generated)
            "Poz No",            # UI col 0
            "Poz Açıklaması",    # UI col 1
            "Ölçü Birimi",       # UI col 2
            "Birim Fiyat",       # UI col 3 (Düzeltildi: 'Birim Fiyatı' yerine 'Birim Fiyat')
            "Metraj",            # UI col 11
            "Yaklaşık İşçilik",  # UI col 12
            "Yaklaşık Maliyet",  # UI col 13
            "Teklif Fiyat",      # UI col 14
            "Teklif İşçilik",    # UI col 15
            "Teklif Maliyet",    # UI col 16 (Düzeltildi: 'Teklif Maliyeti' yerine 'Teklif Maliyet')
            "Gerçek Fiyat",      # UI col 17
            "Gerçek İşçilik",    # UI col 18
            "Gerçek Maliyet",    # UI col 19
        ]

        sheet.append(excel_headers)

        # Header satırına stil uygula
        for col_idx, header_text in enumerate(excel_headers):
            cell = sheet.cell(row=3, column=col_idx + 1)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Verileri topla ve Excel'e yaz
        current_excel_row = 4  # Verilerin yazılmaya başlanacağı satır
        for row_idx in range(table_widget.rowCount()):
            # UI tablosundaki sütun indeksleri ile eşleşen verileri çek
            row_data = [
                row_idx + 1,  # Sıra No (Excel only)
                (table_widget.item(row_idx, 0).text() if table_widget.item(row_idx, 0) else ""),  # Poz No (UI 0)
                (table_widget.item(row_idx, 1).text() if table_widget.item(row_idx, 1) else ""),  # Poz Açıklaması (UI 1)
                (table_widget.item(row_idx, 2).text() if table_widget.item(row_idx, 2) else ""),  # Ölçü Birimi (UI 2)
                (table_widget.item(row_idx, 3).text() if table_widget.item(row_idx, 3) else ""),  # Birim Fiyat (UI 3)
                (table_widget.item(row_idx, 11).text() if table_widget.item(row_idx, 11) else ""), # Metraj (UI 11)
                (table_widget.item(row_idx, 12).text() if table_widget.item(row_idx, 12) else ""), # Yaklaşık İşçilik (UI 12)
                (table_widget.item(row_idx, 13).text() if table_widget.item(row_idx, 13) else ""), # Yaklaşık Maliyet (UI 13)
                (table_widget.item(row_idx, 14).text() if table_widget.item(row_idx, 14) else ""), # Teklif Fiyat (UI 14)
                (table_widget.item(row_idx, 15).text() if table_widget.item(row_idx, 15) else ""), # Teklif İşçilik (UI 15)
                (table_widget.item(row_idx, 16).text() if table_widget.item(row_idx, 16) else ""), # Teklif Maliyet (UI 16)
                (table_widget.item(row_idx, 17).text() if table_widget.item(row_idx, 17) else ""), # Gerçek Fiyat (UI 17)
                (table_widget.item(row_idx, 18).text() if table_widget.item(row_idx, 18) else ""), # Gerçek İşçilik (UI 18)
                (table_widget.item(row_idx, 19).text() if table_widget.item(row_idx, 19) else ""), # Gerçek Maliyet (UI 19)
            ]
            sheet.append(row_data)
            current_excel_row += 1

        # TOPLAM SATIRI EKLE
        toplam_row_excel = current_excel_row

        # "Toplam:" metnini ekleyin (Excel'de 4. sütun olan "Ölçü Birimi" altına denk gelmesi için 4. sütun kullanıldı)
        # veya "Birim Fiyat" sütununun altına "Toplam:" yazmak için 5. sütunu kullanabiliriz.
        # "Metraj" (Excel 6) ile "Yaklaşık İşçilik" (Excel 7) arasına yerleştirelim
        sheet.cell(row=toplam_row_excel, column=5, value="Toplam:").font = Font(
            bold=True
        )

        # Formüller ekleyelim
        data_start_row = 4
        data_end_row = toplam_row_excel - 1

        # Excel sütun indeksleri (1-tabanlı)
        # Yaklaşık Maliyet toplamı (Excel H sütunu, index 8)
        sheet.cell(
            row=toplam_row_excel,
            column=8,
            value=f"=SUM(H{data_start_row}:H{data_end_row})",
        )
        # Teklif Maliyet toplamı (Excel K sütunu, index 11)
        sheet.cell(
            row=toplam_row_excel,
            column=11,
            value=f"=SUM(K{data_start_row}:K{data_end_row})",
        )
        # Gerçek Maliyet toplamı (Excel N sütunu, index 14)
        sheet.cell(
            row=toplam_row_excel,
            column=14,
            value=f"=SUM(N{data_start_row}:N{data_end_row})",
        )

        # Toplam satırı renklendirelim
        total_row_fill = PatternFill(
            start_color="FFD966", end_color="FFD966", fill_type="solid"
        )

        sheet.cell(row=toplam_row_excel, column=5).fill = total_row_fill
        sheet.cell(row=toplam_row_excel, column=8).fill = total_row_fill # Yaklaşık Maliyet toplamı
        sheet.cell(row=toplam_row_excel, column=11).fill = total_row_fill # Teklif Maliyet toplamı
        sheet.cell(row=toplam_row_excel, column=14).fill = total_row_fill # Gerçek Maliyet toplamı

        sheet.cell(row=toplam_row_excel, column=8).font = Font(bold=True)
        sheet.cell(row=toplam_row_excel, column=11).font = Font(bold=True)
        sheet.cell(row=toplam_row_excel, column=14).font = Font(bold=True)


        # Sütun genişliklerini otomatik ayarla
        num_cols = len(excel_headers)

        for col_idx in range(1, num_cols + 1):  # Excel sütun indeksleri 1'den başlar
            column_letter = get_column_letter(col_idx)

            # Özel genişlik ayarları
            if col_idx == 3:  # Poz Açıklaması sütunu (Excel col 3)
                adjusted_width = 40
            elif col_idx == 5: # Birim Fiyat sütunu (Excel col 5)
                adjusted_width = 15
            elif col_idx == 6: # Metraj sütunu (Excel col 6)
                adjusted_width = 15
            else:
                max_length = 0
                # Başlık satırından (row 3) ve veri satırlarından (data_start_row'dan toplam_row_excel'e kadar) değerleri kontrol et
                for row_num in range(3, toplam_row_excel + 1):
                    cell = sheet.cell(row=row_num, column=col_idx)

                    if not isinstance(cell, openpyxl.cell.MergedCell):
                        try:
                            cell_value = (
                                str(cell.value) if cell.value is not None else ""
                            )
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                        except Exception:
                            pass

                adjusted_width = max_length + 2

                # Minimum bir genişlik belirleyebiliriz (isteğe bağlı)
                if adjusted_width < 10:
                    adjusted_width = 10

            sheet.column_dimensions[column_letter].width = adjusted_width

        # Veri hücrelerine stil uygula
        data_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        numeric_alignment_right = Alignment(horizontal="right", vertical="center")

        for r_idx in range(data_start_row, toplam_row_excel + 1):
            for c_idx in range(1, len(excel_headers) + 1): # excel_headers'ın uzunluğu kadar dön
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                cell.font = Font(size=9)

                # Sayısal formatlama gerektiren Excel sütunları (1-tabanlı indeksler)
                # "Birim Fiyat", "Metraj", "Yaklaşık İşçilik", "Yaklaşık Maliyet", "Teklif Fiyat", "Teklif İşçilik", "Teklif Maliyet", "Gerçek Fiyat", "Gerçek İşçilik", "Gerçek Maliyet"
                numeric_cols_for_excel = [5, 6, 7, 8, 9, 10, 11, 12, 13, 14] 
                if c_idx in numeric_cols_for_excel:
                    cell.alignment = numeric_alignment_right
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0.00"
                    elif isinstance(cell.value, str):
                        try:
                            formatted_val = float(cell.value.replace(",", "."))
                            cell.value = formatted_val
                            cell.number_format = "#,##0.00"
                        except ValueError:
                            pass

        workbook.save(excel_filename)

        QMessageBox.information(
            None, "Başarılı", f"Excel dosyası başarıyla oluşturuldu:\n{excel_filename}"
        )
        os.startfile(excel_filename)

    except Exception as e:
        QMessageBox.critical(
            None, "Hata", f"Excel oluşturulurken hata oluştu:\n{str(e)}"
        )
        traceback.print_exc()

