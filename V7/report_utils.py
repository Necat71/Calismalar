# report_utils.py
import os
import sys
import pandas as pd
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from PyQt6.QtCore import QStandardPaths
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill  # Yeni eklendi
from openpyxl.utils import get_column_letter  # Yeni eklendi

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def create_excel_report(
    title,
    content_list,
    suggested_filename="rapor.xlsx",
    parent_window=None,
    custom_dataframes=None,
):  # custom_dataframes eklendi
    """
    Belirtilen başlık ve içerikle bir Excel raporu oluşturur.
    Metinleri bir "Özet" sayfasına, her bir DataFrame'i ise ayrı bir sayfaya yazar.
    Ek olarak, özel DataFrame'ler (örn. Teklif Maliyet Analizi sonuçları) belirtilen sayfalara yazılabilir.
    Kullanıcıya kaydetme konumu sorar ve raporu başarıyla oluşturulursa açar.

    Args:
        title (str): Raporun başlığı. Özet sayfasına yazılır.
        content_list (list): Rapora eklenecek metin (str) veya pandas DataFrame nesnelerinin listesi.
        suggested_filename (str): Kaydetme diyaloğu için önerilen dosya adı.
        parent_window (QWidget): Dosya kaydetme diyaloğunu ortalamak için ana pencere.
        custom_dataframes (dict): Anahtar-değer çiftleri içeren bir sözlük.
                                  Anahtar sayfa adı (str), değer ise pd.DataFrame nesnesidir.
                                  Bu DataFrame'ler belirtilen sayfalara yazılır.
    Returns:
        str: Oluşturulan Excel dosyasının yolu, başarısız olursa None.
    """
    default_dir = QStandardPaths.writableLocation(
        QStandardPaths.StandardLocation.DocumentsLocation
    )

    excel_path, _ = QFileDialog.getSaveFileName(
        parent_window,
        "Excel Raporunu Kaydet",
        os.path.join(default_dir, suggested_filename),
        "Excel Dosyaları (*.xlsx);;Tüm Dosyalar (*.*)",
    )

    if not excel_path:
        print("Excel raporu oluşturma kullanıcı tarafından iptal edildi.")
        return None

    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:

            text_items = [title]
            data_frames = []  # General DataFrames to be written to 'Veri_X' sheets
            for item in content_list:
                if isinstance(item, str):
                    text_items.append(item)
                elif isinstance(item, pd.DataFrame):
                    data_frames.append(item)

            # Write summary text to "Özet" sheet
            summary_df = pd.DataFrame(text_items, columns=["Rapor Özeti"])
            summary_df.to_excel(writer, sheet_name="Özet", index=False)

            worksheet_summary = writer.sheets["Özet"]
            worksheet_summary.column_dimensions["A"].width = 80

            # Write general DataFrames to separate sheets
            for i, df in enumerate(data_frames, start=1):
                sheet_name = f"Veri_{i}"
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                worksheet_data = writer.sheets[sheet_name]
                for col in worksheet_data.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    worksheet_data.column_dimensions[column].width = adjusted_width

            # Handle custom_dataframes for specific reports (like Teklif Maliyet Analizi)
            if custom_dataframes:
                for sheet_name, df_to_write in custom_dataframes.items():
                    df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Apply styling and column width adjustments for these custom sheets
                    worksheet = writer.sheets[sheet_name]

                    # Apply header style if applicable (similar to excel_export_module's header)
                    header_font = Font(bold=True, size=12, color="FFFFFF")
                    header_fill = PatternFill(
                        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
                    )
                    header_alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    thin_border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

                    # Assuming headers are in the first row after writing DataFrame (index=False)
                    for col_idx, cell in enumerate(
                        worksheet[1]
                    ):  # Iterate over cells in the first row
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border

                    # Auto-adjust column width for custom dataframes
                    for col in worksheet.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except (
                                TypeError
                            ):  # Handle cases where cell.value might be None or non-string
                                pass
                        adjusted_width = max_length + 2
                        worksheet.column_dimensions[column].width = adjusted_width

                
                    numeric_cols_for_excel = [
                        "Teklif_Maliyet",
                        "Gerçek_Maliyeti",
                        "teklif_maliyet_farki",
                        "teklif_sapma_yuzdesi",
                    ]  # Add other numeric columns as needed
                    for r_idx in range(
                        2, worksheet.max_row + 1
                    ):  # Start from second row (after headers)
                        for c_idx in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=r_idx, column=c_idx)
                            if (
                                worksheet.cell(row=1, column=c_idx).value
                                in numeric_cols_for_excel
                            ):
                                cell.alignment = Alignment(
                                    horizontal="right", vertical="center"
                                )
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = "#,##0.00"
                                elif isinstance(cell.value, str):
                                    try:
                                        formatted_val = float(
                                            cell.value.replace(",", ".")
                                        )
                                        cell.value = formatted_val
                                        cell.number_format = "#,##0.00"
                                    except ValueError:
                                        pass

                    # Add total row for specific sheets if necessary (e.g., for Teklif Maliyet Analizi)
                    if (
                        sheet_name == "Poz Bazlı Farklar"
                    ):  # Only for this specific sheet
                        total_row_excel = worksheet.max_row + 1
                        worksheet.cell(
                            row=total_row_excel, column=3, value="Toplam:"
                        ).font = Font(
                            bold=True
                        )  # Assuming 'teklif_maliyet_farki' is in column D (index 4) and 'teklif_sapma_yuzdesi' is in E (index 5)

                        # Find column index for 'teklif_maliyet_farki'
                        col_teklif_maliyet_farki = -1
                        for col_idx_h, header_val in enumerate(df_to_write.columns, 1):
                            if header_val == "teklif_maliyet_farki":
                                col_teklif_maliyet_farki = col_idx_h
                                break

                        if col_teklif_maliyet_farki != -1:
                            worksheet.cell(
                                row=total_row_excel,
                                column=col_teklif_maliyet_farki,
                                value=f"=SUM({get_column_letter(col_teklif_maliyet_farki)}2:{get_column_letter(col_teklif_maliyet_farki)}{total_row_excel - 1})",
                            )
                            worksheet.cell(
                                row=total_row_excel, column=col_teklif_maliyet_farki
                            ).font = Font(bold=True)
                            worksheet.cell(
                                row=total_row_excel, column=col_teklif_maliyet_farki
                            ).fill = PatternFill(
                                start_color="FFD966",
                                end_color="FFD966",
                                fill_type="solid",
                            )
                            worksheet.cell(
                                row=total_row_excel, column=col_teklif_maliyet_farki
                            ).number_format = "#,##0.00"

        if sys.platform == "win32":
            os.startfile(excel_path)
        elif sys.platform == "darwin":
            os.system(f'open "{excel_path}"')
        else:
            os.system(f'xdg-open "{excel_path}"')

        return excel_path

    except Exception as e:
        error_message = f"Excel raporu oluşturulurken bir hata oluştu:\n{e}"
        if parent_window:
            QMessageBox.critical(parent_window, "Excel Oluşturma Hatası", error_message)
        else:
            print(error_message)
        return None
